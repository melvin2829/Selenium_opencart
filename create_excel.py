# update_testdata.py
from openpyxl import load_workbook

FILE = "testdata.xlsx"
SHEET = "Sheet1"

# Desired rows (will overwrite existing rows starting after header)
new_rows = [
    ("username", "password", "search_term"),   # header (will keep existing header if present)
    ("Rich@test.com", "test123", "MacBook"),
    ("richy@test.com", "test123", "iPhone")
]

def update_excel(path, sheet_name, rows):
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Clear existing content (keep header row index stable)
    ws.delete_rows(1, ws.max_row)

    # Write new rows
    for r in rows:
        ws.append(r)

    wb.save(path)
    print(f"Updated {path} -> sheet '{sheet_name}' with {len(rows)-1} data rows.")

if __name__ == "__main__":
    update_excel(FILE, SHEET, new_rows)
